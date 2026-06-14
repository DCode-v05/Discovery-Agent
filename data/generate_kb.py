"""Generate a synthetic enterprise knowledge base for the Discovery Agent demo.

Produces a mixed-format document collection under data/kb/ that mentions ~10
systems with deliberately *varying evidence strength* — including one system
mentioned only in passing (to exercise the <70% "flag for review" path) and one
that must be inferred. Reproducible: re-run anywhere with `python data/generate_kb.py`.

Outputs:
  data/kb/architecture_overview.pdf   (PDF, text + a table)
  data/kb/onboarding_wiki.md          (Markdown)
  data/kb/integrations_register.xlsx  (Spreadsheet)
  data/kb/core_systems_inventory.png  (Image — needs vision/OCR)
  data/kb/it_email_thread.txt         (Plain text — passing mentions)
  data/use_cases.json                 (Level 2 automation goals)
"""
from __future__ import annotations

import json
from pathlib import Path

KB = Path(__file__).resolve().parent / "kb"
KB.mkdir(parents=True, exist_ok=True)
DATA = Path(__file__).resolve().parent


def make_pdf() -> None:
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.lib import colors

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(KB / "architecture_overview.pdf"), pagesize=LETTER)
    body = styles["BodyText"]
    h1, h2 = styles["Heading1"], styles["Heading2"]
    story = [
        Paragraph("Acme Corp — Enterprise Systems Architecture Overview", h1),
        Paragraph("Prepared by Platform Engineering. Internal. Q2 review.", body),
        Spacer(1, 12),
        Paragraph("1. Customer & Revenue Systems", h2),
        Paragraph(
            "Our system of record for customers and pipeline is <b>Salesforce</b> (CRM). "
            "It manages Account, Contact, Opportunity, and Quote objects and supports the "
            "lead-to-cash process. Integrations authenticate using OAuth 2.0 (connected app).",
            body),
        Paragraph(
            "Finance runs on <b>NetSuite</b> (ERP). It owns Sales Order, Invoice, and Vendor "
            "Bill records and supports invoicing and revenue recognition. The NetSuite REST "
            "API uses token-based authentication (TBA).",
            body),
        Spacer(1, 8),
        Paragraph("2. Procurement & Data", h2),
        Paragraph(
            "Procurement is handled by <b>Coupa</b>. It owns Purchase Requisition and Purchase "
            "Order entities and supports the procure-to-pay process; the Coupa API uses an API key.",
            body),
        Paragraph(
            "Analytics is centralised in <b>Snowflake</b>, our cloud data warehouse. It ingests "
            "data from finance and product for reporting. Access uses key-pair authentication.",
            body),
        Spacer(1, 8),
        Paragraph("3. Identity", h2),
        Paragraph(
            "Single sign-on is provided by <b>Okta</b>. All employee application access is "
            "federated through Okta via SAML. User and Group are the core entities.",
            body),
        Spacer(1, 12),
        Paragraph("Summary table", h2),
    ]
    table_data = [
        ["System", "Category", "Auth", "Criticality"],
        ["Salesforce", "CRM", "OAuth 2.0", "Critical"],
        ["NetSuite", "ERP", "Token (TBA)", "Critical"],
        ["Coupa", "Procurement", "API Key", "High"],
        ["Snowflake", "Data Warehouse", "Key-pair", "High"],
        ["Okta", "Identity / SSO", "SAML", "Critical"],
    ]
    tbl = Table(table_data, hAlign="LEFT")
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#eef2f7")]),
    ]))
    story.append(tbl)
    doc.build(story)
    print("wrote architecture_overview.pdf")


def make_markdown() -> None:
    md = """# New Engineer Onboarding — Tools & Access

Welcome to Acme! Here's how to get access to the systems you'll use day to day.

## Day 1 — Identity
Your accounts are provisioned automatically when HR adds you in **Workday** (our HRIS,
the system of record for Employee and Position records). Within an hour you'll be able
to log in via **Okta** single sign-on. Okta federates access to every internal app using
SAML, so you only sign in once.

## Day 1 — Communication
Join the engineering channels in **Slack**. Most async discussion and incident comms
happen there.

## Week 1 — CRM
If you're customer-facing, request a **Salesforce** seat through the Okta dashboard.
Salesforce is where we track Accounts and Opportunities. Access uses your SSO identity.

> Note: provisioning a new hire touches Workday -> Okta -> Salesforce in sequence.
"""
    (KB / "onboarding_wiki.md").write_text(md, encoding="utf-8")
    print("wrote onboarding_wiki.md")


def make_spreadsheet() -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Integrations"
    headers = ["System", "Category", "Auth Method", "Key Entities", "Existing Integrations", "Criticality"]
    ws.append(headers)
    rows = [
        ["Salesforce", "CRM", "OAuth 2.0", "Account, Opportunity, Quote", "NetSuite (orders)", "Critical"],
        ["NetSuite", "ERP", "Token (TBA)", "Sales Order, Invoice, Bill", "Salesforce (orders)", "Critical"],
        ["Coupa", "Procurement", "API Key", "Requisition, Purchase Order", "(none)", "High"],
        ["Snowflake", "Data Warehouse", "Key-pair", "Warehouse tables", "(none)", "High"],
        ["Stripe", "Payments", "API Key (Bearer)", "Charge, Customer, Payout", "(none)", "High"],
        ["Workday", "HRIS", "OAuth 2.0", "Employee, Position", "Okta (SCIM)", "Critical"],
    ]
    for r in rows:
        ws.append(r)
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for col, width in zip("ABCDEF", [16, 16, 18, 30, 26, 12]):
        ws.column_dimensions[col].width = width
    wb.save(KB / "integrations_register.xlsx")
    print("wrote integrations_register.xlsx")


def make_image() -> None:
    from PIL import Image, ImageDraw, ImageFont

    W, H = 1000, 760
    img = Image.new("RGB", (W, H), "white")
    d = ImageDraw.Draw(img)
    try:
        title_font = ImageFont.truetype("arial.ttf", 34)
        body_font = ImageFont.truetype("arial.ttf", 22)
    except OSError:
        title_font = ImageFont.load_default()
        body_font = ImageFont.load_default()

    d.rectangle([0, 0, W, 70], fill="#1F3A5F")
    d.text((30, 20), "Acme Corp — Core Systems Inventory (screenshot)", fill="white", font=title_font)
    lines = [
        "1. Salesforce        CRM             OAuth 2.0        Critical",
        "2. NetSuite          ERP             Token (TBA)      Critical",
        "3. Coupa             Procurement     API Key          High",
        "4. Snowflake         Data Warehouse  Key-pair         High",
        "5. Stripe            Payments        API Key          High",
        "6. Workday           HRIS            OAuth 2.0        Critical",
        "7. Okta              Identity/SSO    SAML             Critical",
    ]
    y = 110
    for line in lines:
        d.text((40, y), line, fill="black", font=body_font)
        y += 48
    d.text((40, y + 20), "Source: IT asset register export, 2026-Q2", fill="#555555", font=body_font)
    img.save(KB / "core_systems_inventory.png")
    print("wrote core_systems_inventory.png")


def make_email() -> None:
    txt = """From: priya.n@acme.example
To: it-ops@acme.example
Subject: Re: backlog cleanup before the integration project

Hi all,

Quick notes from today's sync so we don't lose them:

- Reminder that we still log customer support tickets in Zendesk. Nobody owns the
  migration yet, so for now treat it as out of scope but don't delete the account.
- The marketing team wants their lead-capture automation platform connected to the CRM
  so new leads route to Salesforce automatically. They'll share the platform details
  later this week.
- Finance confirmed Stripe payouts should land in the warehouse for the revenue dashboard.

Will follow up after the architecture review.

Thanks,
Priya
"""
    (KB / "it_email_thread.txt").write_text(txt, encoding="utf-8")
    print("wrote it_email_thread.txt")


def make_use_cases() -> None:
    use_cases = [
        {"id": "uc1", "name": "Quote-to-Cash sync",
         "description": "When a Salesforce Opportunity is Closed-Won, create the matching Sales Order and Invoice in NetSuite automatically.",
         "frequency_per_year": 5200, "criticality": "critical"},
        {"id": "uc2", "name": "Procure-to-Pay automation",
         "description": "Sync approved Purchase Orders from Coupa into NetSuite as Vendor Bills for payment.",
         "frequency_per_year": 1200, "criticality": "high"},
        {"id": "uc3", "name": "New-hire provisioning",
         "description": "When HR adds an employee in Workday, provision their Okta identity and a Salesforce seat.",
         "frequency_per_year": 300, "criticality": "high"},
        {"id": "uc4", "name": "Revenue reporting pipeline",
         "description": "Load NetSuite financials and Stripe payouts into Snowflake nightly for the revenue dashboard.",
         "frequency_per_year": 365, "criticality": "high"},
        {"id": "uc5", "name": "Support ticket enrichment",
         "description": "Push resolved Zendesk tickets into Salesforce as Cases linked to the Account.",
         "frequency_per_year": 8000, "criticality": "medium"},
        {"id": "uc6", "name": "Marketing lead routing",
         "description": "Route new leads from the marketing automation platform into Salesforce as Leads.",
         "frequency_per_year": 15000, "criticality": "medium"},
    ]
    (DATA / "use_cases.json").write_text(json.dumps(use_cases, indent=2), encoding="utf-8")
    print("wrote use_cases.json")


def main() -> None:
    make_pdf()
    make_markdown()
    make_spreadsheet()
    make_image()
    make_email()
    make_use_cases()
    print("\nKnowledge base generated at:", KB)


if __name__ == "__main__":
    main()
